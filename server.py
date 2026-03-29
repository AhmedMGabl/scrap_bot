import json, os, re, subprocess, sys, threading, time
from datetime import datetime
from pathlib import Path
from flask import Flask, Response, jsonify, request, stream_with_context
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

app = Flask(__name__)
BASE_DIR = Path(__file__).parent
DATA_FILE = BASE_DIR / "data" / "teams.json"
SCHEDULE_FILE = BASE_DIR / "data" / "schedule.json"
CONFIG_FILE = BASE_DIR / "data" / "config.json"
DATA_FILE.parent.mkdir(exist_ok=True)

DEFAULT_CONFIG = {
    "active_webhook_url": "https://oapi.dingtalk.com/robot/send?access_token=28bc378d0fc40e94d1ae14f3223373c8d6fe6654e6595dd4ff6a138ecc3de0a3",
    "dingtalk_groups": [
        {"name": "Hany TEST", "url": "https://oapi.dingtalk.com/robot/send?access_token=28bc378d0fc40e94d1ae14f3223373c8d6fe6654e6595dd4ff6a138ecc3de0a3"}
    ],
    "crm_username": "51Hany",
    "crm_password": "b%7DWWtm",
    "ams_username": "51hany",
    "ams_password": "Hyoussef@51",
}

def _load_config():
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text())
        except Exception:
            pass
    return dict(DEFAULT_CONFIG)

def _save_config(data):
    CONFIG_FILE.write_text(json.dumps(data, indent=2))

# ── scheduler ─────────────────────────────────────────────────────────────────

scheduler = BackgroundScheduler(timezone="Africa/Cairo")
_schedule_lock = threading.Lock()


def _load_schedule():
    if SCHEDULE_FILE.exists():
        try:
            return json.loads(SCHEDULE_FILE.read_text())
        except Exception:
            pass
    return {"jobs": []}


def _save_schedule(data):
    SCHEDULE_FILE.write_text(json.dumps(data, indent=2))


_PHASE_SCRIPTS = {
    "crm":  "Scripts/scrape_crm_report.py",
    "ams":  "Scripts/scrape_iur_report.py",
    "cm":   "generate_cm_report.py",
    "ea":   "generate_ea_report.py",
    "send": "run_daily_report.py",
}

def _run_pipeline_job(job_id, phases, test_mode, channel):
    """Called by APScheduler — runs each phase script directly."""
    print(f"[scheduler] Job {job_id} triggered at {datetime.now().isoformat()}")
    extra = ["--test"] if test_mode else []
    for phase in (phases or list(_PHASE_SCRIPTS.keys())):
        script = _PHASE_SCRIPTS.get(phase)
        if not script:
            continue
        args = extra.copy()
        if phase == "send":
            args.append("--send-only")
        cmd = [sys.executable, str(BASE_DIR / script)] + args
        try:
            result = subprocess.run(cmd, cwd=str(BASE_DIR), timeout=3600)
            print(f"[scheduler] Phase {phase}: exit {result.returncode}")
        except subprocess.TimeoutExpired:
            print(f"[scheduler] Phase {phase} timed out")
        except Exception as e:
            print(f"[scheduler] Phase {phase} error: {e}")


def _register_jobs():
    """Register all enabled jobs from schedule.json with APScheduler."""
    scheduler.remove_all_jobs()
    data = _load_schedule()
    for job in data.get("jobs", []):
        if not job.get("enabled", True):
            continue
        try:
            hour, minute = job["hour"], job.get("minute", 0)
            # days_of_week: 0=Mon … 4=Fri  (APScheduler uses mon/tue/wed/thu/fri)
            days = job.get("days", "mon-fri")
            scheduler.add_job(
                _run_pipeline_job,
                CronTrigger(hour=hour, minute=minute, day_of_week=days,
                            timezone="Africa/Cairo"),
                id=job["id"],
                args=[job["id"], job.get("phases", ["crm","ams","cm","ea","send"]),
                      job.get("test", True), job.get("channel", "dingtalk")],
                replace_existing=True,
            )
            print(f"[scheduler] Registered job {job['id']} at {hour:02d}:{minute:02d} ({days})")
        except Exception as e:
            print(f"[scheduler] Failed to register job {job.get('id')}: {e}")


_register_jobs()
scheduler.start()

# ── helpers ──────────────────────────────────────────────────────────────────

def _run_script(script_name, extra_args=None):
    """Run a project script and yield SSE log lines."""
    script = str(BASE_DIR / script_name)
    cmd = [sys.executable, script] + (extra_args or [])
    proc = subprocess.Popen(
        cmd, cwd=str(BASE_DIR),
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        text=True, bufsize=1
    )
    for line in proc.stdout:
        line = line.rstrip()
        if line:
            yield f"data: {json.dumps({'type': 'log', 'text': line})}\n\n"
    proc.wait()
    rc = proc.returncode
    status = "ok" if rc == 0 else f"error (exit {rc})"
    yield f"data: {json.dumps({'type': 'log', 'text': f'[{script_name}] finished: {status}'})}\n\n"

# ── API routes ────────────────────────────────────────────────────────────────

@app.route("/api/ping")
def ping():
    return jsonify({"ok": True})


@app.route("/api/health")
def health():
    output_dir = BASE_DIR / "Output"
    input_dir = BASE_DIR / "Input"
    rawdata = input_dir / "rawdata.xlsx"
    ts_file = input_dir / "Team Structure.xlsx"
    crm_cookie = (BASE_DIR / "Scripts" / "crm_cookies.json").exists()

    sheet1_rows = None
    sheet2_rows = None
    rawdata_mtime = None

    if rawdata.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(rawdata), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            if len(sheet_names) > 0:
                ws1 = wb[sheet_names[0]]
                sheet1_rows = ws1.max_row - 1  # subtract header
            if len(sheet_names) > 1:
                ws2 = wb[sheet_names[1]]
                sheet2_rows = ws2.max_row - 1
            wb.close()
        except Exception:
            pass
        mtime = rawdata.stat().st_mtime
        rawdata_mtime = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")

    logs = sorted((output_dir / "logs").glob("*.log"), reverse=True) if (output_dir / "logs").exists() else []
    last_log = str(logs[0].name) if logs else None

    return jsonify({
        "ok": True,
        "last_run": last_log,
        "output_files": len(list(output_dir.glob("*.*"))) if output_dir.exists() else 0,
        "sheet1_rows": sheet1_rows,
        "sheet2_rows": sheet2_rows,
        "rawdata_mtime": rawdata_mtime,
        "ts_ok": ts_file.exists(),
        "crm_cookie": crm_cookie,
    })


@app.route("/api/run")
def run_pipeline():
    phases = request.args.get("phases", "crm,ams,cm,ea,send").split(",")
    test_mode = request.args.get("test", "true").lower() != "false"
    channel = request.args.get("channel", "dingtalk")
    group = request.args.get("group", "test")

    def generate():
        yield f"data: {json.dumps({'type': 'log', 'text': f'Starting phases: {phases} | test={test_mode} | channel={channel} | group={group}'})}\n\n"
        for phase in phases:
            yield f"data: {json.dumps({'type': 'phase', 'phase': phase, 'state': 'running'})}\n\n"
            try:
                extra = []
                if test_mode:
                    extra.append("--test")
                if phase == "crm":
                    yield from _run_script("Scripts/scrape_crm_report.py", extra)
                elif phase == "ams":
                    yield from _run_script("Scripts/scrape_iur_report.py", extra)
                elif phase == "cm":
                    yield from _run_script("generate_cm_report.py", extra)
                elif phase == "ea":
                    yield from _run_script("generate_ea_report.py", extra)
                elif phase == "send":
                    yield from _run_script("run_daily_report.py", extra + ["--send-only"])
                yield f"data: {json.dumps({'type': 'phase', 'phase': phase, 'state': 'done'})}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type': 'log', 'text': f'[error] {phase}: {e}'})}\n\n"
                yield f"data: {json.dumps({'type': 'phase', 'phase': phase, 'state': 'error'})}\n\n"
        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Schedule CRUD ─────────────────────────────────────────────────────────────

@app.route("/api/schedule", methods=["GET"])
def get_schedule():
    data = _load_schedule()
    # Annotate each job with next_run from APScheduler
    for job in data.get("jobs", []):
        apjob = scheduler.get_job(job["id"])
        if apjob and apjob.next_run_time:
            job["next_run"] = apjob.next_run_time.strftime("%Y-%m-%d %H:%M %Z")
        else:
            job["next_run"] = None
    return jsonify(data)


@app.route("/api/schedule", methods=["POST", "OPTIONS"])
def add_schedule():
    if request.method == "OPTIONS":
        return "", 204
    body = request.get_json(force=True)
    hour = int(body.get("hour", 15))
    minute = int(body.get("minute", 0))
    days = body.get("days", "mon-fri")
    phases = body.get("phases", ["crm", "ams", "cm", "ea", "send"])
    test_mode = body.get("test", True)
    channel = body.get("channel", "dingtalk")
    label = body.get("label", f"{hour:02d}:{minute:02d}")
    enabled = body.get("enabled", True)

    with _schedule_lock:
        data = _load_schedule()
        job_id = f"job_{int(time.time())}"
        new_job = {
            "id": job_id,
            "label": label,
            "hour": hour,
            "minute": minute,
            "days": days,
            "phases": phases,
            "test": test_mode,
            "channel": channel,
            "enabled": enabled,
        }
        data["jobs"].append(new_job)
        _save_schedule(data)
        _register_jobs()

    return jsonify({"ok": True, "id": job_id})


@app.route("/api/schedule/<job_id>", methods=["DELETE", "OPTIONS"])
def delete_schedule(job_id):
    if request.method == "OPTIONS":
        return "", 204
    with _schedule_lock:
        data = _load_schedule()
        data["jobs"] = [j for j in data["jobs"] if j["id"] != job_id]
        _save_schedule(data)
        _register_jobs()
    return jsonify({"ok": True})


@app.route("/api/schedule/<job_id>/toggle", methods=["POST", "OPTIONS"])
def toggle_schedule(job_id):
    if request.method == "OPTIONS":
        return "", 204
    with _schedule_lock:
        data = _load_schedule()
        for job in data["jobs"]:
            if job["id"] == job_id:
                job["enabled"] = not job.get("enabled", True)
                break
        _save_schedule(data)
        _register_jobs()
    return jsonify({"ok": True})


# ── Other routes ──────────────────────────────────────────────────────────────

@app.route("/api/scheduler/setup", methods=["POST", "OPTIONS"])
def scheduler_setup():
    if request.method == "OPTIONS": return "", 204
    try:
        subprocess.Popen([sys.executable, str(BASE_DIR / "setup_schedule.py")], cwd=str(BASE_DIR))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/scheduler/remove", methods=["POST", "OPTIONS"])
def scheduler_remove():
    if request.method == "OPTIONS": return "", 204
    try:
        subprocess.Popen([sys.executable, str(BASE_DIR / "remove_schedule.py")], cwd=str(BASE_DIR))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/send/retry", methods=["POST", "OPTIONS"])
def send_retry():
    if request.method == "OPTIONS": return "", 204
    try:
        subprocess.Popen([sys.executable, str(BASE_DIR / "lark_sender.py")], cwd=str(BASE_DIR))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/cookies/clear", methods=["DELETE", "OPTIONS"])
def cookies_clear():
    if request.method == "OPTIONS": return "", 204
    cleared = []
    for f in (BASE_DIR / "Scripts").glob("*.json"):
        if "cookie" in f.name.lower():
            f.unlink()
            cleared.append(f.name)
    return jsonify({"ok": True, "cleared": cleared})


@app.route("/api/upload", methods=["POST", "OPTIONS"])
def upload():
    if request.method == "OPTIONS": return "", 204
    saved = []
    for key, f in request.files.items():
        dest = BASE_DIR / "Input" / f.filename
        dest.parent.mkdir(exist_ok=True)
        f.save(str(dest))
        saved.append(f.filename)
    return jsonify({"ok": True, "saved": saved})


@app.route("/api/teams", methods=["GET"])
def get_teams():
    if DATA_FILE.exists():
        return jsonify(json.loads(DATA_FILE.read_text()))
    return jsonify({})


@app.route("/api/teams", methods=["POST", "OPTIONS"])
def save_teams():
    if request.method == "OPTIONS": return "", 204
    DATA_FILE.write_text(json.dumps(request.get_json(force=True), indent=2))
    return jsonify({"ok": True})


@app.route("/api/config", methods=["GET"])
def get_config():
    return jsonify(_load_config())


@app.route("/api/config", methods=["POST", "OPTIONS"])
def save_config_route():
    if request.method == "OPTIONS": return "", 204
    _save_config(request.get_json(force=True))
    return jsonify({"ok": True})


@app.route("/api/logs", methods=["GET"])
def list_logs():
    log_dir = BASE_DIR / "Output" / "logs"
    if not log_dir.exists():
        return jsonify([])
    logs = []
    for f in sorted(log_dir.glob("*.log"), reverse=True)[:50]:
        try:
            stat = f.stat()
            logs.append({
                "name": f.name,
                "size": stat.st_size,
                "mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
        except Exception:
            pass
    return jsonify(logs)


@app.route("/api/logs/<filename>", methods=["GET"])
def get_log_file(filename):
    if not re.match(r'^[\w\-\.]+\.log$', filename):
        return jsonify({"error": "invalid filename"}), 400
    log_file = BASE_DIR / "Output" / "logs" / filename
    if not log_file.exists():
        return jsonify({"error": "not found"}), 404
    return log_file.read_text(errors="replace"), 200, {"Content-Type": "text/plain; charset=utf-8"}


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, threaded=True)
