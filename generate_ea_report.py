import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

# Import HTML generation functions from CM report
from generate_cm_report import (
    generate_html_individual_report,
    generate_html_separate_teams_report,
    generate_html_bottom20_report
)
from html_report_generator import generate_html_team_report




def get_color_for_value(value, thresholds):
    """Get Excel color based on value and thresholds [red_max, orange_max]"""
    if value < thresholds[0]:
        return 'FCA5A5'
    elif value < thresholds[1]:
        return 'FED7AA'
    else:
        return 'D1FAE5'



# Excel report generation functions
def generate_excel_reports(merged_df, output_dir):
    """Generate all Excel reports"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    
    # Individual Report
    sorted_df = merged_df.sort_values('Total Duration (Min)', ascending=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'EA Individual'
    ws.append(['Team', 'EA', 'Total Calls', 'Total Eff. Calls', 'Total Duration (Min)', 'Avg Call Time/Min'])
    for i, r in sorted_df.iterrows():
        ws.append([r['Team'], r['CRM'], int(r['Total Calls']), int(r['Total Eff. Calls']), int(round(r['Total Duration (Min)'])), round(r['Avg Call Time/Min'], 2)])
    wb.save(output_dir + '/EA_Individual_Report.xlsx')
    print(f'Excel individual report saved')
    
    # Team Summary
    ts = merged_df.groupby('Team').agg({'CRM': 'count', 'Total Eff. Calls': 'sum', 'Total Duration (Min)': 'sum'}).reset_index()
    ts['Avg Eff. Calls'] = (ts['Total Eff. Calls'] / ts['CRM']).round(1)
    # Avg Call Time/Min per team = Total Duration / Members
    ts['Avg Call Time/Min'] = (ts['Total Duration (Min)'] / ts['CRM']).fillna(0).round(0).astype(int)
    ts.columns = ['Team', 'Members', 'Total Eff. Calls', 'Total Duration (Min)', 'Avg Eff. Calls', 'Avg Call Time/Min']
    ts = ts[['Team', 'Members', 'Total Eff. Calls', 'Total Duration (Min)', 'Avg Eff. Calls', 'Avg Call Time/Min']]
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = 'EA Team Summary'
    ws2.append(list(ts.columns))
    for i, r in ts.iterrows():
        ws2.append(list(r))
    wb2.save(output_dir + '/EA_Team_Summary.xlsx')
    print(f'Excel team summary saved')
    
    # Separate Teams
    wb3 = openpyxl.Workbook()
    wb3.remove(wb3.active)
    for team in sorted(merged_df['Team'].unique()):
        td = merged_df[merged_df['Team'] == team].sort_values('Total Duration (Min)', ascending=False)
        ws3 = wb3.create_sheet(title=str(team)[:31])
        ws3.append(['EA', 'Total Calls', 'Total Eff. Calls', 'Total Duration (Min)', 'Avg Call Time/Min'])
        for i, r in td.iterrows():
            ws3.append([r['CRM'], int(r['Total Calls']), int(r['Total Eff. Calls']), int(round(r['Total Duration (Min)'])), round(r['Avg Call Time/Min'], 2)])
        ws3.append([])
        total_members = len(td)
        total_tc = int(td['Total Calls'].sum()); total_tec = int(td['Total Eff. Calls'].sum()); total_dur = int(round(td['Total Duration (Min)'].sum()))
        team_avg_ct = int(round(total_dur / total_members)) if total_members > 0 else 0
        ws3.append(['TOTAL', total_tc, total_tec, total_dur, team_avg_ct])
        ws3.append(['AVERAGE', round(td['Total Calls'].mean(), 1), round(td['Total Eff. Calls'].mean(), 1), int(round(td['Total Duration (Min)'].mean())), round(td['Avg Call Time/Min'].mean(), 2)])
    wb3.save(output_dir + '/EA_Separate_Teams.xlsx')
    print(f'Excel separate teams report saved')
    
    # Bottom 20
    b20 = merged_df.sort_values('Total Duration (Min)', ascending=True).head(20)
    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    ws4.title = 'Bottom 20'
    ws4.append(['Rank', 'Team', 'EA', 'Total Calls', 'Total Eff. Calls', 'Total Duration (Min)', 'Avg Call Time/Min'])
    for rank, (i, r) in enumerate(b20.iterrows(), 1):
        ws4.append([rank, r['Team'], r['CRM'], int(r['Total Calls']), int(r['Total Eff. Calls']), int(round(r['Total Duration (Min)'])), round(r['Avg Call Time/Min'], 2)])
    wb4.save(output_dir + '/EA_Bottom20.xlsx')
    print(f'Excel bottom 20 report saved')


def read_duration_data_from_file(file_path, sheet_name='Sheet1'):
    """Read duration data from a single file"""
    print(f"  Reading: {os.path.basename(file_path)}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    # Find headers (row 3)
    headers = []
    for col in range(1, 16):
        cell_value = ws.cell(row=3, column=col).value
        if cell_value:
            headers.append(cell_value)
    
    # Find required column indices
    col_indices = {}
    for i, header in enumerate(headers):
        if header == 'SC':  # EA uses 'SC' instead of 'CRM'
            col_indices['crm'] = i
        elif header == 'Total number of calls':
            col_indices['total_calls'] = i
        elif header == 'Total valid calls':  # EA uses this instead of 'Effective Calls'
            col_indices['eff_calls'] = i
        elif header == 'Total effective call time/Minute':  # EA column name
            col_indices['duration'] = i
        elif header == 'Average call time/Minute':
            col_indices['avg_call'] = i
        # Note: EA data doesn't have class consumption, will default to 0
    
    # Read data starting from row 4
    data = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        crm_value = row[col_indices['crm']]
        # Skip if CRM is None, 'Total', 'In total', or contains Chinese characters (group headers)
        if crm_value and isinstance(crm_value, (int, str)):
            if isinstance(crm_value, str):
                if crm_value in ['Total', 'In total', 'SC'] or '小组' in crm_value or '组' in crm_value:
                    continue
            
            total_calls_val = 0
            if col_indices.get('total_calls') is not None and row[col_indices['total_calls']]:
                total_calls_val = float(row[col_indices['total_calls']])
            data.append({
                'CRM': str(crm_value),  # Convert to string for consistency
                'Total Calls': total_calls_val,
                'Total Eff. Calls': float(row[col_indices['eff_calls']]) if row[col_indices['eff_calls']] else 0,
                'Total Duration (Min)': float(row[col_indices['duration']]) if row[col_indices['duration']] else 0,
                'Avg Call Time/Min': float(row[col_indices['avg_call']]) if row[col_indices['avg_call']] else 0,
            })
    
    df = pd.DataFrame(data)
    print(f"    Records: {len(df)}")
    return df


def aggregate_monthly_data(monthly_files):
    """
    Aggregate duration data from multiple monthly files
    
    Args:
        monthly_files: List of file paths (can be single file for testing)
    
    Returns:
        Aggregated DataFrame with cumulative metrics
    """
    print("Reading monthly duration data...")
    all_data = []
    
    for file_path in monthly_files:
        if os.path.exists(file_path):
            df = read_duration_data_from_file(file_path)
            all_data.append(df)
        else:
            print(f"  WARNING: File not found - {file_path}")
    
    if not all_data:
        raise FileNotFoundError("No valid duration data files found!")
    
    # Combine all monthly data
    combined_df = pd.concat(all_data, ignore_index=True)
    print(f"\nCombined records from all months: {len(combined_df)}")
    
    # Aggregate by CRM - sum calls/duration, use MEAN for avg (NO RECALCULATION!)
    aggregated_df = combined_df.groupby('CRM').agg({
        'Total Calls': 'sum',
        'Total Eff. Calls': 'sum',
        'Total Duration (Min)': 'sum',
        'Avg Call Time/Min': 'mean'
    }).reset_index()
    
    print(f"Aggregated EA employees: {len(aggregated_df)}")
    return aggregated_df


def read_ea_structure(file_path):
    """Read EA structure from Team Structure.xlsx"""
    print("\nReading EA structure...")
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    # Try 'EA' first, fall back to first sheet
    ea_sheet = 'EA' if 'EA' in sheet_names else sheet_names[0]
    if ea_sheet != 'EA':
        print(f"  Warning: 'EA' sheet not found, using '{ea_sheet}'")
    df = pd.read_excel(file_path, sheet_name=ea_sheet)
    df.columns = ['Team', 'CRM']
    print(f"EA Structure: {len(df)} members across {df['Team'].nunique()} teams")
    return df


def merge_ea_data(duration_df, ea_structure_df):
    """Merge aggregated duration data with EA structure"""
    print("\nMerging EA data...")
    
    # Merge on CRM
    merged_df = pd.merge(ea_structure_df, duration_df, on='CRM', how='left')
    
    # Fill NaN values with 0 for employees with no activity
    merged_df['Total Calls'] = merged_df['Total Calls'].fillna(0).astype(int)
    merged_df['Total Eff. Calls'] = merged_df['Total Eff. Calls'].fillna(0).astype(int)
    merged_df['Total Duration (Min)'] = merged_df['Total Duration (Min)'].fillna(0)
    merged_df['Avg Call Time/Min'] = merged_df['Avg Call Time/Min'].fillna(0)
    merged_df['Classes Completed'] = 0  # EA doesn't track classes
    
    print(f"Merged data: {len(merged_df)} EA members")
    return merged_df


def main():
    # Get script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # File paths
    ea_structure_file = os.path.join(script_dir, 'Input', 'Team Structure.xlsx')
    output_dir = os.path.join(script_dir, 'Output')
    
    # ========================================================================
    # CONFIGURE YOUR MONTHLY DATA FILES HERE
    # ========================================================================
    
    # EA Data File (November 2025 - January 2026 aggregated)
    monthly_files = [
        os.path.join(script_dir, 'Input', 'EA_rawdata_Nov_Jan.xlsx')
    ]
    
    # Note: If you have separate monthly files, you can specify them as a list:
    # monthly_files = [
    #     os.path.join(script_dir, 'Input', 'rawdata_november.xlsx'),
    #     os.path.join(script_dir, 'Input', 'rawdata_december.xlsx'),
    #     os.path.join(script_dir, 'Input', 'rawdata_january.xlsx'),
    # ]
    
    # ========================================================================
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    print("="*70)
    print("EA REPORT GENERATION (November - January Aggregated)")
    print("="*70)
    
    # Read EA structure
    ea_structure_df = read_ea_structure(ea_structure_file)
    
    # Read and aggregate duration data from all months
    duration_df = aggregate_monthly_data(monthly_files)
    
    # Merge data
    merged_df = merge_ea_data(duration_df, ea_structure_df)
    
    # Generate reports
    print("\n" + "="*70)
    print("GENERATING HTML REPORTS")
    print("="*70)
    
    individual_report_file = os.path.join(output_dir, 'EA_Individual_Report.html')
    team_summary_file = os.path.join(output_dir, 'EA_Team_Summary.html')
    separate_teams_file = os.path.join(output_dir, 'EA_Separate_Teams.html')
    bottom20_file = os.path.join(output_dir, 'EA_Bottom20.html')
    
    # Generate all HTML reports
    generate_html_individual_report(merged_df, individual_report_file)
    generate_html_team_report(merged_df, team_summary_file)
    generate_html_separate_teams_report(merged_df, separate_teams_file)
    generate_html_bottom20_report(merged_df, bottom20_file)
    
    # Generate all Excel reports
    generate_excel_reports(merged_df, output_dir)
    
    print("\n" + "="*70)
    print("EA REPORTS GENERATED SUCCESSFULLY!")
    print("="*70)
    print(f"\nIndividual Report: {os.path.abspath(individual_report_file)}")
    print(f"Team Summary:      {os.path.abspath(team_summary_file)}")
    print(f"Separate Teams:    {os.path.abspath(separate_teams_file)}")
    print(f"Bottom 20:         {os.path.abspath(bottom20_file)}")
    
    # Display summary statistics
    print("\n" + "="*70)
    print("EA DATA SUMMARY")
    print("="*70)
    print(f"\nTotal EA Members: {len(merged_df)}")
    print(f"Total Teams: {merged_df['Team'].nunique()}")
    print(f"\nOverall Metrics (Nov-Jan Combined):")
    print(f"  Total Eff. Calls: {merged_df['Total Eff. Calls'].sum():,.0f}")
    print(f"  Total Duration: {merged_df['Total Duration (Min)'].sum():,.0f} minutes")
    print(f"  Average Call Time/Min: {merged_df['Avg Call Time/Min'].mean():.2f}")
    
    print("\nTop 5 Teams by Total Duration:")
    top_teams = merged_df.groupby('Team')['Total Duration (Min)'].sum().sort_values(ascending=False).head()
    for team, duration in top_teams.items():
        print(f"  {team}: {duration:,.0f} min")


if __name__ == "__main__":
    main()
