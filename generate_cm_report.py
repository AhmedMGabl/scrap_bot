import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), "Scripts"))
from html_report_generator import generate_html_team_report
import json
import shutil
from datetime import datetime
import time


# ===== CRM SCRAPING FUNCTIONS =====

def load_crm_config():
    """Load CRM configuration from environment variables or crm_config.json"""

    # Try environment variables first (Docker/production)
    if os.getenv('CRM_USERNAME'):
        print("Loading CRM config from environment variables...")
        config = {
            'crm_username': os.getenv('CRM_USERNAME'),
            'crm_password': os.getenv('CRM_PASSWORD'),
            'crm_url': os.getenv('CRM_URL'),
            'headless': os.getenv('HEADLESS', 'true').lower() == 'true',
            'timeout': int(os.getenv('TIMEOUT', '60'))
        }
        print(f"OK Loaded CRM config for: {config['crm_username']} (from env)")
        return config

    # Fallback to JSON file (local development)
    print("Loading CRM config from crm_config.json...")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_file = os.path.join(script_dir, 'crm_config.json')

    if not os.path.exists(config_file):
        print("ERROR: crm_config.json not found!")
        print("Please create it with your CRM credentials.")
        raise FileNotFoundError(f"CRM config file not found: {config_file}")

    with open(config_file, 'r', encoding='utf-8') as f:
        config = json.load(f)

    required_fields = ['crm_username', 'crm_password', 'crm_url']
    missing = [f for f in required_fields if f not in config]
    if missing:
        raise ValueError(f"Missing fields: {missing}")

    config.setdefault('headless', False)
    config.setdefault('timeout', 30)
    print(f"OK Loaded CRM config for: {config['crm_username']} (from file)")
    return config


def scrape_crm_data(config):
    """Scrape data from CRM using Playwright"""
    from playwright.sync_api import sync_playwright
    import time as _t
    print("Launching browser...")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, 'Output')
    os.makedirs(output_dir, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=config['headless'])
        page = browser.new_page(viewport={'width': 1400, 'height': 900})
        page.set_default_timeout(config['timeout'] * 1000)
        page.on('dialog', lambda dialog: dialog.accept())

        # Step 1: Go directly to the sideline login page (no link clicking needed)
        print("Step 1: Navigating to CRM sideline login...")
        page.goto('https://crm.51talk.com/admin/admin_login.php?login_employee_type=sideline&redirect_uri=')
        page.wait_for_load_state('domcontentloaded')
        _t.sleep(2)

        # Step 2: Fill credentials (field name is 'user_name', NOT 'username')
        print("Step 2: Entering credentials...")
        page.fill('input[name="user_name"]', config['crm_username'])
        page.fill('input[name="password"]', config['crm_password'])

        # Step 3: Submit
        print("Step 3: Submitting login...")
        page.click('input[name="Submit"]')
        page.wait_for_load_state('networkidle')
        _t.sleep(2)
        print("Logged in successfully!")

        # Step 4: Navigate to report URL
        print(f"Step 4: Navigating to: {config['crm_url']}")
        page.goto(config['crm_url'])
        page.wait_for_load_state('networkidle')
        _t.sleep(3)

        shot = os.path.join(output_dir, f"crm_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        page.screenshot(path=shot, full_page=True)
        print(f"OK Screenshot: {shot}")

        # Step 5: Extract table via JS
        print("Step 5: Extracting table data...")
        raw_rows = page.evaluate("""() => {
            const rows = Array.from(document.querySelectorAll('table tr'));
            return rows.map(tr =>
                Array.from(tr.querySelectorAll('td, th')).map(td => td.innerText.trim())
            ).filter(r => r.length > 1);
        }""")

        browser.close()

        if not raw_rows:
            print("WARNING: No table data found")
            return []

        headers = raw_rows[0]
        crm_data = [dict(zip(headers, row)) for row in raw_rows[1:] if len(row) == len(headers)]
        print(f"OK Extracted {len(crm_data)} rows with {len(headers)} columns")
        return crm_data


def format_crm_data_for_rawdata(crm_data):
    """Format scraped CRM data"""
    print("\nFormatting data...")
    df = pd.DataFrame(crm_data)
    print(f"Columns: {list(df.columns)}")
    print("NOTE: Column mapping may need adjustment after first run")
    return df


def update_rawdata_tab1(formatted_data, rawdata_file):
    """Update Tab 1 in rawdata.xlsx"""
    print("\nUpdating rawdata.xlsx...")
    backup_dir = os.path.join(os.path.dirname(rawdata_file), 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    backup_file = os.path.join(backup_dir, f"rawdata_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    if os.path.exists(rawdata_file):
        shutil.copy2(rawdata_file, backup_file)
        print(f"OK Backup: {backup_file}")
        backups = sorted([f for f in os.listdir(backup_dir) if f.startswith('rawdata_backup_')])
        for old_backup in backups[:-5]:
            os.remove(os.path.join(backup_dir, old_backup))
    wb = openpyxl.load_workbook(rawdata_file)
    ws = wb['1']
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)
    for idx, row in formatted_data.iterrows():
        row_num = idx + 4
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
    wb.save(rawdata_file)
    wb.close()
    print(f"OK Updated with {len(formatted_data)} rows")


def scrape_and_update_rawdata():
    """Main CRM scraping orchestration"""
    print("="*60)
    print("PHASE 1: Scraping CRM Data")
    print("="*60)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    rawdata_file = os.path.join(script_dir, 'Input', 'rawdata.xlsx')
    try:
        config = load_crm_config()
        crm_data = scrape_crm_data(config)
        formatted_data = format_crm_data_for_rawdata(crm_data)
        update_rawdata_tab1(formatted_data, rawdata_file)
        print("=" * 60)
        print("OK CRM SCRAPING COMPLETED SUCCESSFULLY")
        print("=" * 60)
    except Exception as e:
        print(f"  First attempt failed: {e}, retrying...")
        try:
            config = load_crm_config()
            crm_data = scrape_crm_data(config)
            formatted_data = format_crm_data_for_rawdata(crm_data)
            update_rawdata_tab1(formatted_data, rawdata_file)
            print("=" * 60)
            print("OK CRM SCRAPING COMPLETED SUCCESSFULLY")
            print("=" * 60)
        except Exception as e:
            print("=" * 60)
            print(f"ERROR CRM SCRAPING FAILED: {e}")
            print("Proceeding with existing rawdata.xlsx...")
            print("=" * 60)
            raise

# ===== END OF CRM SCRAPING FUNCTIONS =====


def generate_screenshots(html_files, output_dir):
    """Generate screenshots of HTML reports using Playwright"""
    from playwright.sync_api import sync_playwright
    import time
    
    print("\nGenerating screenshots...")
    
    with sync_playwright() as p:
        # Launch browser
        browser = p.chromium.launch()
        page = browser.new_page(viewport={'width': 1400, 'height': 900})
        
        for html_file in html_files:
            # Get the base name without extension
            base_name = os.path.splitext(os.path.basename(html_file))[0]
            screenshot_file = os.path.join(output_dir, f'{base_name}.png')
            
            # Navigate to the HTML file
            file_url = f'file:///{os.path.abspath(html_file).replace(chr(92), "/")}'
            page.goto(file_url, wait_until='networkidle')
            page.screenshot(path=screenshot_file, full_page=True)
            print(f"Screenshot saved: {screenshot_file}")
        
        browser.close()
    
    print("All screenshots generated successfully!")


def generate_html_individual_report(merged_df, output_file):
    """Generate HTML individual member report"""
    
    # Sort by Total Duration (Min) descending (best performers first)
    sorted_df = merged_df.sort_values('Total Duration (Min)', ascending=False)
    
    # Calculate min/max for dynamic color scaling
    duration_min = merged_df['Total Duration (Min)'].min()
    duration_max = merged_df['Total Duration (Min)'].max()
    avg_call_min = merged_df['Avg Call Time/Min'].min()
    avg_call_max = merged_df['Avg Call Time/Min'].max()

    def get_color_class(value, min_val, max_val):
        """Excel-style color scaling: min=red, mid=yellow, max=green"""
        if max_val == min_val:
            return 'bg-medium'
        
        # Normalize value to 0-1 range
        normalized = (value - min_val) / (max_val - min_val)
        
        # Excel 3-color scale: 
        # 0.0 = red (bottom), 0.5 = yellow (middle), 1.0 = green (top)
        if normalized < 0.5:
            # Lower half: red to yellow
            if normalized < 0.17:
                return 'bg-very-low'
            elif normalized < 0.33:
                return 'bg-low'
            else:
                return 'bg-medium-low'
        else:
            # Upper half: yellow to green
            if normalized < 0.67:
                return 'bg-medium'
            elif normalized < 0.83:
                return 'bg-medium-high'
            else:
                return 'bg-high'

    
    # Generate HTML
    html = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Individual CM Report</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            background: #f5f7fa;
            padding: 30px;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            padding: 30px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }
        
        .header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 30px;
        }
        
        h1 {
            font-size: 24px;
            font-weight: 600;
            color: #1a1a1a;
        }
        
        .download-btn {
            background: #f5f7fa;
            border: 1px solid #e1e4e8;
            padding: 8px 16px;
            border-radius: 6px;
            color: #586069;
            font-size: 14px;
            cursor: pointer;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .download-btn:hover {
            background: #e9ecef;
        }
        
        table {
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
        }
        
        thead th {
            background: #f5f7fa;
            color: #6b7280;
            font-weight: 600;
            font-size: 13px;
            text-align: left;
            padding: 14px 16px;
            border-bottom: 2px solid #e5e7eb;
            position: sticky;
            top: 0;
        }
        
        tbody td {
            padding: 12px 16px;
            border-bottom: 1px solid #f3f4f6;
            font-size: 14px;
            color: #1f2937;
        }
        
        tbody tr:hover {
            background: #fafbfc;
        }
        
        .badge {
            display: inline-block;
            padding: 6px 14px;
            border-radius: 20px;
            font-weight: 500;
            font-size: 13px;
            text-align: center;
            min-width: 60px;
        }
        
        .bg-high {
            background: #63BE7B;
            color: #000;
        }
        
        .bg-medium-high {
            background: #9FD899;
            color: #000;
        }
        
        .bg-medium {
            background: #C6E5B5;
            color: #000;
        }
        
        .bg-medium-low {
            background: #FFEB84;
            color: #000;
        }
        
        .bg-low {
            background: #FCAA75;
            color: #000;
        }
        
        .bg-very-low {
            background: #F8696B;
            color: #000;
        }
        
        .text-center {
            text-align: center !important;
        }
        
        .team-row {
            background: #f9fafb;
            font-weight: 600;
        }
        
        @media print {
            body {
                padding: 0;
                background: white;
            }
            
            .container {
                box-shadow: none;
                padding: 20px;
            }
            
            .download-btn {
                display: none;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Individual CM Performance Report</h1>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Team</th>
                    <th>Name</th>
                    <th class="text-center">Total Calls</th>
                    <th class="text-center">Total Eff. Calls</th>
                    <th class="text-center">Total Duration (Min)</th>
                    <th class="text-center">Avg Call Time/Min</th>
                    <th class="text-center">Classes Completed</th>
                    
                </tr>
            </thead>
            <tbody>
"""
    
    # Add data rows
    for idx, row in sorted_df.iterrows():
        duration_color = get_color_class(row['Total Duration (Min)'], duration_min, duration_max)
        avg_call_color = get_color_class(row['Avg Call Time/Min'], avg_call_min, avg_call_max)
        
        html += f"""
                <tr>
                    <td>{row['Team']}</td>
                    <td>{row['Name']}</td>
                    <td class="text-center">{row['Total Calls']}</td>
                    <td class="text-center">{row['Total Eff. Calls']}</td>
                    <td class="text-center"><span class="badge {duration_color}">{int(round(row['Total Duration (Min)']))}</span></td>
                    <td class="text-center"><span class="badge {avg_call_color}">{int(round(row['Avg Call Time/Min']))}</span></td>
                    <td class="text-center">{row['Classes Completed']}</td>
                </tr>
"""
    
    html += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"HTML individual report saved to: {output_file}")




def generate_html_separate_teams_report(merged_df, output_file):
    sorted_df = merged_df.sort_values(['Team', 'Total Duration (Min)'], ascending=[True, False])
    
    def get_color_class_relative(value, values_in_team, col_name):
        """Excel-style color scaling within each team"""
        if len(values_in_team) == 0 or value == 0:
            return 'bg-very-low'
        
        valid_values = [v for v in values_in_team if v > 0]
        if len(valid_values) == 0:
            return 'bg-very-low'
        
        min_val = min(valid_values)
        max_val = max(valid_values)
        
        if max_val == min_val:
            return 'bg-medium'
        
        # Normalize value to 0-1 range
        normalized = (value - min_val) / (max_val - min_val)
        
        # Excel 3-color scale
        if normalized < 0.5:
            if normalized < 0.17:
                return 'bg-very-low'
            elif normalized < 0.33:
                return 'bg-low'
            else:
                return 'bg-medium-low'
        else:
            if normalized < 0.67:
                return 'bg-medium'
            elif normalized < 0.83:
                return 'bg-medium-high'
            else:
                return 'bg-high'
        
        sorted_values = sorted([v for v in values_in_team if v > 0], reverse=True)
        if len(sorted_values) == 0:
            return 'bg-very-low'
        
        # Find percentile
        if value in sorted_values:
            rank = sorted_values.index(value) + 1
            percentile = (len(sorted_values) - rank + 1) / len(sorted_values)
            
            # 6-level gradient for better visualization
            if percentile >= 0.83:
                return 'bg-high'
            elif percentile >= 0.67:
                return 'bg-medium-high'
            elif percentile >= 0.50:
                return 'bg-medium'
            elif percentile >= 0.33:
                return 'bg-medium-low'
            elif percentile >= 0.17:
                return 'bg-low'
            else:
                return 'bg-very-low'
        return 'bg-very-low'
    
    html = '<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Separate Teams</title><style>* { margin: 0; padding: 0; } body { font-family: sans-serif; background: #f5f7fa; padding: 30px; } .container { max-width: 1400px; margin: 0 auto; } .header { background: white; border-radius: 12px; padding: 30px; margin-bottom: 20px; } h1 { font-size: 24px; } .team-section { background: white; border-radius: 12px; padding: 25px; margin-bottom: 20px; } .team-header { font-size: 18px; font-weight: 600; margin-bottom: 15px; border-bottom: 2px solid #e5e7eb; padding-bottom: 10px; } table { width: 100%; } thead th { background: #f5f7fa; padding: 12px; border-bottom: 2px solid #e5e7eb; } tbody td { padding: 12px; border-bottom: 1px solid #f3f4f6; } .badge { display: inline-block; padding: 6px 14px; border-radius: 20px; font-weight: 500; min-width: 60px; text-align: center; } .bg-high { background: #63BE7B; color: #000; } .bg-medium-high { background: #9FD899; color: #000; } .bg-medium { background: #C6E5B5; color: #000; } .bg-medium-low { background: #FFEB84; color: #000; } .bg-low { background: #FCAA75; color: #000; } .bg-very-low { background: #F8696B; color: #000; } .total-row { background: #dbeafe; font-weight: 700; border-top: 2px solid #3b82f6 !important; } .total-row td { padding: 14px 12px !important; color: #1e3a5f; } .average-row { background: #f3e8ff; font-weight: 600; border-top: 1px solid #a855f7 !important; } .average-row td { padding: 14px 12px !important; color: #4a1d7a; } .text-center { text-align: center; }</style></head><body><div class="container"><div class="header"><h1>CM Report by Separate Teams (Relative Performance)</h1></div>'
    
    teams = sorted_df['Team'].unique()
    for team in teams:
        team_data = sorted_df[sorted_df['Team'] == team]
        
        # Get all values for this team for relative comparison
        duration_values = team_data['Total Duration (Min)'].tolist()
        avg_call_values = team_data['Avg Call Time/Min'].tolist()
        
        html += f'<div class="team-section"><div class="team-header">{team} ({len(team_data)} members)</div><table><thead><tr><th>Name</th><th class="text-center">Total Calls</th><th class="text-center">Total Eff. Calls</th><th class="text-center">Total Duration (Min)</th><th class="text-center">Avg Call Time/Min</th><th class="text-center">Classes Completed</th></tr></thead><tbody>'
        
        for _, row in team_data.iterrows():
            dc = get_color_class_relative(row['Total Duration (Min)'], duration_values, 'Total Duration (Min)')
            ac = get_color_class_relative(row['Avg Call Time/Min'], avg_call_values, 'Avg Call Time/Min')
            dval = int(round(row['Total Duration (Min)'])) if row['Total Duration (Min)'] > 0 else 0
            aval = int(round(row['Avg Call Time/Min'])) if row['Avg Call Time/Min'] > 0 else 0
            html += f'<tr><td>{row["Name"]}</td><td class="text-center">{row["Total Calls"]}</td><td class="text-center">{row["Total Eff. Calls"]}</td><td class="text-center"><span class="badge {dc}">{dval}</span></td><td class="text-center"><span class="badge {ac}">{aval}</span></td><td class="text-center">{row["Classes Completed"]}</td></tr>'
        
        # Calculate totals and averages for summary rows
        total_calls = int(team_data['Total Calls'].sum())
        total_eff_calls = int(team_data['Total Eff. Calls'].sum())
        total_duration = int(round(team_data['Total Duration (Min)'].sum()))
        total_classes = int(team_data['Classes Completed'].sum())
        
        total_members = len(team_data)
        avg_calls = round(total_calls / total_members, 1) if total_members > 0 else 0
        avg_eff_calls = round(total_eff_calls / total_members, 1) if total_members > 0 else 0
        avg_duration = int(round(team_data['Total Duration (Min)'].mean()))
        # AVERAGE Avg Call Time/Min = mean of individual values
        avg_call_time = int(round(team_data['Avg Call Time/Min'].apply(lambda x: int(round(x)) if x > 0 else 0).mean()))
        avg_classes = int(round(team_data['Classes Completed'].mean()))
        
        # TOTAL row: Avg Call Time/Min = Total Duration / Total Eff. Calls
        team_avg_call_time = int(round(total_duration / total_eff_calls)) if total_eff_calls > 0 else 0
        
        # Add TOTAL row
        html += f'<tr class="total-row"><td><strong>TOTAL</strong></td><td class="text-center"><strong>{total_calls}</strong></td><td class="text-center"><strong>{total_eff_calls}</strong></td><td class="text-center"><strong>{total_duration}</strong></td><td class="text-center"><strong>{team_avg_call_time}</strong></td><td class="text-center"><strong>{total_classes}</strong></td></tr>'
        
        # Add AVERAGE row
        html += f'<tr class="average-row"><td><strong>AVERAGE</strong></td><td class="text-center">{avg_calls}</td><td class="text-center">{avg_eff_calls}</td><td class="text-center">{avg_duration}</td><td class="text-center">{avg_call_time}</td><td class="text-center">{avg_classes}</td></tr>'
        
        html += '</tbody></table></div>'
    
    html += '</div></body></html>'
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"HTML separate teams report saved to: {output_file}")

def generate_html_bottom20_report(merged_df, output_file):
    bottom20_df = merged_df.sort_values('Total Duration (Min)', ascending=True).head(20)
    
    # Calculate min/max for dynamic color scaling  
    duration_min = bottom20_df['Total Duration (Min)'].min()
    duration_max = bottom20_df['Total Duration (Min)'].max()
    avg_call_min = bottom20_df['Avg Call Time/Min'].min()
    avg_call_max = bottom20_df['Avg Call Time/Min'].max()
    
    def get_color_class(value, min_val, max_val):
        """Excel-style color scaling: min=red, mid=yellow, max=green"""
        if max_val == min_val:
            return 'bg-medium'
    
        # Normalize value to 0-1 range
        normalized = (value - min_val) / (max_val - min_val)
    
        # Excel 3-color scale: 
        # 0.0 = red (bottom)
        # 0.5 = yellow (middle) 
        # 1.0 = green (top)
        if normalized < 0.5:
            # Lower half: red to yellow gradient
            # Map 0-0.5 to our 3 red/orange/yellow levels
            if normalized < 0.17:
                return 'bg-very-low'  # Pure red
            elif normalized < 0.33:
                return 'bg-low'  # Orange
            else:
                return 'bg-medium-low'  # Yellow
        else:
            # Upper half: yellow to green gradient
            # Map 0.5-1.0 to our 3 yellow/light green/dark green levels
            if normalized < 0.67:
                return 'bg-medium'  # Light green
            elif normalized < 0.83:
                return 'bg-medium-high'  # Medium green
            else:
                return 'bg-high'  # Dark green
    html = '<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Bottom 20</title><style>* { margin: 0; padding: 0; } body { font-family: sans-serif; background: #f5f7fa; padding: 30px; } .container { max-width: 1400px; margin: 0 auto; background: white; border-radius: 12px; padding: 30px; } h1 { font-size: 24px; margin-bottom: 5px; } .subtitle { color: #6b7280; font-size: 14px; margin-bottom: 30px; } table { width: 100%; } thead th { background: #f5f7fa; padding: 14px; border-bottom: 2px solid #e5e7eb; } tbody td { padding: 12px; border-bottom: 1px solid #f3f4f6; } .rank { font-weight: 600; color: #991b1b; } .badge { display: inline-block; padding: 6px 14px; border-radius: 20px; font-weight: 500; min-width: 60px; text-align: center; } .bg-high { background: #63BE7B; color: #000; } .bg-medium-high { background: #9FD899; color: #000; } .bg-medium { background: #C6E5B5; color: #000; } .bg-medium-low { background: #FFEB84; color: #000; } .bg-low { background: #FCAA75; color: #000; } .bg-very-low { background: #F8696B; color: #000; } .summary-row { background: #f9fafb; font-weight: 600; border-top: 2px solid #e5e7eb !important; } .summary-row td { padding: 14px 12px !important; } .text-center { text-align: center; }</style></head><body><div class="container"><h1>Bottom 20 CM Performance Report</h1><div class="subtitle">Ranked by Total Duration (Min) (Lowest to Highest)</div><table><thead><tr><th class="text-center">Rank</th><th>Team</th><th>Name</th><th class="text-center">Total Calls</th><th class="text-center">Total Eff. Calls</th><th class="text-center">Total Duration (Min)</th><th class="text-center">Avg Call Time/Min</th><th class="text-center">Classes Completed</th></tr></thead><tbody>'
    for rank, (_, row) in enumerate(bottom20_df.iterrows(), 1):
        dc = get_color_class(row['Total Duration (Min)'], duration_min, duration_max)
        ac = get_color_class(row['Avg Call Time/Min'], avg_call_min, avg_call_max)
        dval = int(round(row['Total Duration (Min)'])) if row['Total Duration (Min)'] > 0 else 0
        aval = int(round(row['Avg Call Time/Min'])) if row['Avg Call Time/Min'] > 0 else 0
        html += f'<tr><td class="text-center rank">{rank}</td><td>{row["Team"]}</td><td>{row["Name"]}</td><td class="text-center">{row["Total Calls"]}</td><td class="text-center">{row["Total Eff. Calls"]}</td><td class="text-center"><span class="badge {dc}">{dval}</span></td><td class="text-center"><span class="badge {ac}">{aval}</span></td><td class="text-center">{row["Classes Completed"]}</td></tr>'
    html += '</tbody></table></div></body></html>'
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"HTML bottom 20 report saved to: {output_file}")


def read_duration_data(file_path):
    """Read duration data from Tab 1"""
    try:
        df = pd.read_excel(file_path, sheet_name='1')
        print(f"Columns found: {list(df.columns)[:5]}...")  # Print first 5 columns
        
        # Clean the dataframe - remove rows that are all NaN or have NaN in key columns
        required_cols = ['SC', 'Total valid calls']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValueError(f"Required columns missing: {missing}. Found: {list(df.columns)}")
        df = df.dropna(subset=['SC', 'Total valid calls'], how='any')
        
        # Filter to only rows with actual data (serial number is numeric)
        serial_col = 'Serial' if 'Serial' in df.columns else 'Serial number'
        if serial_col in df.columns:
            df = df[df[serial_col].apply(lambda x: isinstance(x, (int, float)) and x > 0)]
        
        return df
    except Exception as e:
        print(f"Error reading duration data: {e}")
        return pd.DataFrame()

def read_iur_data(file_path):
    """Read IUR classes data from Tab 2"""
    df = pd.read_excel(file_path, sheet_name='2')
    # Filter out subtotal rows (useraccount1 == '小计') and grand total row (NaN)
    df = df[df['useraccount1'].notna() & (df['useraccount1'] != '小计') & (df['useraccount1'].str.lower() != 'total')].copy()
    # Keep only relevant columns
    df = df[['useraccount1', 'Class completed']].copy()
    df.columns = ['CRM', 'Classes Completed']
    return df

def read_cm_structure(file_path):
    """Read CM structure data from CM sheet"""
    df = pd.read_excel(file_path, sheet_name='CM')
    # Rename 'CM Team' to 'Team' for consistency
    if 'CM Team' in df.columns:
        df = df.rename(columns={'CM Team': 'Team'})
    return df

def normalize_name(name):
    """Normalize name for matching"""
    if pd.isna(name):
        return ''
    name = str(name).strip().lower()
    # Remove common prefixes
    for prefix in ['eglp-', 'egss-', 'jolp-', 'joss-', '51']:
        if name.startswith(prefix):
            name = name[len(prefix):]
    return name

def merge_all_data(duration_df, iur_df, cm_structure_df):
    """Merge all data sources"""
    
    # Prepare duration data - exact case-insensitive matching on SC column
    duration_df = duration_df.copy()
    duration_df['CRM_key'] = duration_df['SC'].str.strip().str.lower()
    
    # Prepare IUR data
    iur_df = iur_df.copy()
    iur_df['CRM_key'] = iur_df['CRM'].str.strip().str.lower()
    
    # Prepare CM structure
    cm_structure_df = cm_structure_df.copy()
    cm_structure_df['CRM_key'] = cm_structure_df['CRM'].str.strip().str.lower()
    
    # Ensure Total number of calls column exists
    if 'Total number of calls' not in duration_df.columns:
        duration_df['Total number of calls'] = 0
    
    dur_cols = ['CRM_key', 'Total number of calls', 'Total valid calls', 'Total effective call time/Minute', 'Average call time/Minute']
    
    # Merge duration with CM structure
    merged = cm_structure_df.merge(
        duration_df[dur_cols],
        on='CRM_key',
        how='left'
    )
    
    # Merge with IUR data
    merged = merged.merge(
        iur_df[['CRM_key', 'Classes Completed']],
        on='CRM_key',
        how='left'
    )
    
    # Fill NaN values
    merged['Total number of calls'] = merged['Total number of calls'].fillna(0).astype(int)
    merged['Total valid calls'] = merged['Total valid calls'].fillna(0).astype(int)
    merged['Total effective call time/Minute'] = merged['Total effective call time/Minute'].fillna(0)
    merged['Average call time/Minute'] = merged['Average call time/Minute'].fillna(0)
    merged['Classes Completed'] = merged['Classes Completed'].fillna(0).astype(int)
    # Rename columns for clarity
    merged = merged.rename(columns={
        'CRM': 'Name',
        'Total number of calls': 'Total Calls',
        'Total valid calls': 'Total Eff. Calls',
        'Total effective call time/Minute': 'Total Duration (Min)',
        'Average call time/Minute': 'Avg Call Time/Min',
    })
    
    return merged[['Team', 'Name', 'Total Calls', 'Total Eff. Calls', 'Total Duration (Min)', 'Avg Call Time/Min', 'Classes Completed']]

def get_color_for_value(value, thresholds):
    """Get color based on value and thresholds"""
    # thresholds = [red_max, orange_max, yellow_max] - anything above yellow_max is green
    if value <= thresholds[0]:
        return 'E74C3C'  # Red
    elif value <= thresholds[1]:
        return 'E67E22'  # Orange
    elif value <= thresholds[2]:
        return 'F39C12'  # Yellow
    else:
        return '27AE60'  # Green

def create_individual_report(merged_df, output_file):
    """Create individual member report with color coding"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Individual Report"
    
    # Headers
    headers = ['Team', 'Name', 'Total Calls', 'Total Eff. Calls', 'Total Duration (Min)', 'Avg Call Time/Min', 'Classes Completed']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Define thresholds for color coding
    # These can be adjusted based on your performance criteria
    duration_thresholds = [400, 500, 600]  # Red <= 400, Orange <= 500, Yellow <= 600, Green > 600
    avg_call_thresholds = [35, 40, 45]     # Red <= 35, Orange <= 40, Yellow <= 45, Green > 45
    
    # Add data rows
    for idx, row in merged_df.iterrows():
        ws.append([
            row['Team'],
            row['Name'],
            row['Total Calls'],
            row['Total Eff. Calls'],
            row['Total Duration (Min)'],
            row['Avg Call Time/Min'],
            row['Classes Completed'],
        ])
        
        current_row = ws.max_row
        
        # Apply color coding to Total Duration column (column 5)
        duration_cell = ws.cell(row=current_row, column=5)
        duration_value = row['Total Duration (Min)']
        duration_color = get_color_for_value(duration_value, duration_thresholds)
        duration_cell.fill = PatternFill(start_color=duration_color, end_color=duration_color, fill_type='solid')
        duration_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply color coding to Avg Call Time column (column 6)
        avg_call_cell = ws.cell(row=current_row, column=6)
        avg_call_value = row['Avg Call Time/Min']
        avg_call_color = get_color_for_value(avg_call_value, avg_call_thresholds)
        avg_call_cell.fill = PatternFill(start_color=avg_call_color, end_color=avg_call_color, fill_type='solid')
        avg_call_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Center align all cells
        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    column_widths = [15, 25, 12, 14, 20, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    wb.save(output_file)
    print(f"Individual report saved to: {output_file}")

def create_team_summary(merged_df, output_file):
    """Create team summary report"""
    
    # Group by team and calculate aggregates
    team_summary = merged_df.groupby('Team').agg({
        'Name': 'count',
        'Total Eff. Calls': 'sum',
        'Total Duration (Min)': 'sum',
        'Classes Completed': 'sum',
    }).reset_index()
    
    # Calculate average effective calls per team member
    team_summary['Avg Eff. Calls'] = (team_summary['Total Eff. Calls'] / team_summary['Name']).round(1)
    
    # Calculate Avg Call Time/Min as Total Duration / Members
    avg_call_by_team = merged_df.groupby('Team')['Avg Call Time/Min'].apply(
        lambda vals: sum(int(round(v)) for v in vals)
    ).rename('Avg Call Time/Min')
    team_summary = team_summary.merge(avg_call_by_team, on='Team')
    
    # Rename columns
    team_summary.columns = ['Team', 'Members', 'Total Eff. Calls', 'Total Duration (Min)',
                            'Classes Completed', 'Avg Eff. Calls', 'Avg Call Time/Min']
    
    # Reorder columns
    team_summary = team_summary[['Team', 'Members', 'Total Eff. Calls', 'Total Duration (Min)', 
                                 'Avg Eff. Calls', 'Avg Call Time/Min', 'Classes Completed']]
    
    # Sort by team name
    team_summary = team_summary.sort_values('Team')
    
    # Add total row
    total_eff_calls = team_summary['Total Eff. Calls'].sum()
    total_duration = team_summary['Total Duration (Min)'].sum()
    total_members = team_summary['Members'].sum()
    
    total_row = {
        'Team': 'TOTAL',
        'Members': total_members,
        'Total Eff. Calls': total_eff_calls,
        'Total Duration (Min)': total_duration,
        'Avg Eff. Calls': round(total_eff_calls / total_members, 1) if total_members > 0 else 0,
        'Avg Call Time/Min': int(total_duration / total_members) if total_members > 0 else 0,
        'Classes Completed': team_summary['Classes Completed'].sum(),
    }
    
    team_summary = pd.concat([team_summary, pd.DataFrame([total_row])], ignore_index=True)
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Team Summary"
    
    # Write headers
    headers = list(team_summary.columns)
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Define thresholds for team summary color coding
    duration_thresholds = [450, 550, 650]
    avg_call_thresholds = [35, 40, 50]
    
    # Add data rows
    for idx, row in team_summary.iterrows():
        ws.append([
            row['Team'],
            row['Members'],
            row['Total Eff. Calls'],
            row['Total Duration (Min)'],
            row['Avg Eff. Calls'],
            row['Avg Call Time/Min'],
            row['Classes Completed'],
        ])
        
        current_row = ws.max_row
        
        # Skip color coding for TOTAL row
        if row['Team'] != 'TOTAL':
            # Apply color coding to Total Duration column (column 4)
            duration_cell = ws.cell(row=current_row, column=4)
            duration_value = row['Total Duration (Min)']
            duration_color = get_color_for_value(duration_value, duration_thresholds)
            duration_cell.fill = PatternFill(start_color=duration_color, end_color=duration_color, fill_type='solid')
            
            # Apply color coding to Avg Call Time column (column 6)
            avg_call_cell = ws.cell(row=current_row, column=6)
            avg_call_value = row['Avg Call Time/Min']
            avg_call_color = get_color_for_value(avg_call_value, avg_call_thresholds)
            avg_call_cell.fill = PatternFill(start_color=avg_call_color, end_color=avg_call_color, fill_type='solid')
        else:
            # Bold the TOTAL row
            for col in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col).font = Font(bold=True, size=11)
        
        # Center align all cells
        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    column_widths = [15, 12, 16, 20, 15, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    wb.save(output_file)
    print(f"Team summary report saved to: {output_file}")

def main():
    # Get the directory where this script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # File paths - updated for local environment
    rawdata_file = os.path.join(script_dir, 'Input', 'rawdata.xlsx')
    cm_structure_file = os.path.join(script_dir, 'Input', 'Team Structure.xlsx')
    output_dir = os.path.join(script_dir, 'Output')

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # PHASE 1: CRM Scraping (NEW)
    print("\n" + "="*60)

    try:
        scrape_and_update_rawdata()
    except Exception as e:
        print(f"CRM scraping failed: {e}")
        print("Continuing with existing data...")
    
    print("\n" + "="*60)

    print("PHASE 2: Generating Reports")
    print("="*60)


    print("Reading data files...")
    
    # Read all data
    duration_df = read_duration_data(rawdata_file)
    print(f"Duration data: {len(duration_df)} records")
    
    iur_df = read_iur_data(rawdata_file)
    print(f"IUR data: {len(iur_df)} records")
    
    cm_structure_df = read_cm_structure(cm_structure_file)
    print(f"CM Structure: {len(cm_structure_df)} members")
    
    # Merge all data
    print("\nMerging data...")
    merged_df = merge_all_data(duration_df, iur_df, cm_structure_df)
    print(f"Merged data: {len(merged_df)} records")
    
    # Create reports
    print("\nGenerating reports...")
    
    individual_report_file = os.path.join(output_dir, 'CM_Individual_Report.html')
    html_report_file = os.path.join(output_dir, 'CM_Team_Summary.html')
    
    # Generate HTML reports
    generate_html_individual_report(merged_df, individual_report_file)
    generate_html_team_report(merged_df, html_report_file)

    # Generate additional reports
    separate_teams_file = os.path.join(output_dir, 'CM_Separate_Teams.html')
    bottom20_file = os.path.join(output_dir, 'CM_Bottom20.html')
    
    generate_html_separate_teams_report(merged_df, separate_teams_file)
    generate_html_bottom20_report(merged_df, bottom20_file)
    
    # Generate screenshots of all reports
    html_files = [individual_report_file, html_report_file, separate_teams_file, bottom20_file]
    generate_screenshots(html_files, output_dir)
    
    print("\n" + "="*60)
    print("Reports generated successfully!")
    print("="*60)
    print(f"Individual Report (HTML): {os.path.abspath(individual_report_file)}")
    print(f"Team Summary (HTML): {os.path.abspath(html_report_file)}")
    print(f"Separate Teams Report (HTML): {os.path.abspath(separate_teams_file)}")
    print(f"Bottom 20 Report (HTML): {os.path.abspath(bottom20_file)}")
    
    # Display sample of merged data
    print("\nSample of merged data:")
    print(merged_df.head(10))
    
    print("\nTeam summary preview:")
    team_preview = merged_df.groupby('Team').agg({
        'Name': 'count',
        'Total Duration (Min)': 'sum',
        'Classes Completed': 'sum'
    })
    print(team_preview)

if __name__ == "__main__":
    main()
